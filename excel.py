import openpyxl 
from openpyxl.styles import Font, Color, colors, Border, Side, Alignment, PatternFill
import n_queen


wb = openpyxl.Workbook()
sheet=wb.active
def xl_create(n,queen_lst,m):

    for i in range(1,1000):
        sheet.row_dimensions[i].height=50

    black_fill = PatternFill("solid", fgColor="000000")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    black_font = Font(size=20, bold=True, color="000000")
    white_font = Font(size=20, bold=True, color="FFFFFF")
    
    for k in range(0,m):
        temp = int(m ** 0.5)
        for i in range(1,n+1):
            for j in range(1,n+1):
                if queen_lst[k][i-1]==j:
                    sheet.cell(row=i+(int(k/(temp+1)))*(n+1),column=j+(k%(temp+1))*(n+1)).value="\u2655"
                    #此处代码过长使用“\”来换行
                    sheet.cell(row=i+(int(k/(temp+1)))*(n+1),column=j+(k%(temp+1))*(n+1)).alignment =\
                    Alignment(horizontal='center',vertical='center',wrap_text=True)
                if (i+j)%2==0:
                    sheet.cell(row=i+(int(k/(temp+1)))*(n+1),column=j+(k%(temp+1))*(n+1)).fill = white_fill
                    sheet.cell(row=i+(int(k/(temp+1)))*(n+1),column=j+(k%(temp+1))*(n+1)).font = black_font
                if (i+j)%2==1:
                    sheet.cell(row=i+(int(k/(temp+1)))*(n+1),column=j+(k%(temp+1))*(n+1)).fill = black_fill
                    sheet.cell(row=i+(int(k/(temp+1)))*(n+1),column=j+(k%(temp+1))*(n+1)).font = white_font
        if (k>900): break; 


if __name__ == "__main__":
    num = eval(input("请输入皇后的个数："))
    queen_list =n_queen.solve(num)
    xl_create(num,queen_list,len(queen_list))
    wb.save("queen.xlsx")
    print("共{:d}种可能".format(len(queen_list)))